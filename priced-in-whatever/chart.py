import pandas as pd
import openpyxl
import plotly.graph_objects as go

# create a list of all dates for the javascript 'Chart.js'

def dateList(filename):
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb["Sheet1"]
    date = []
    for i in range(2, sheet1.max_row + 1):
        date.append(sheet1[f"A{i}"].value)
    return date


# create a list of all ratios for the javascript 'Chart.js'

def openList(filename):
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb["Sheet1"]
    price = []
    for i in range(2, sheet1.max_row + 1):
        price.append(sheet1[f"K{i}"].value)
    return price

def highList(filename):
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb["Sheet1"]
    price = []
    for i in range(2, sheet1.max_row + 1):
        price.append(sheet1[f"L{i}"].value)
    return price

def lowList(filename):
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb["Sheet1"]
    price = []
    for i in range(2, sheet1.max_row + 1):
        price.append(sheet1[f"M{i}"].value)
    return price

def closeList(filename):
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb["Sheet1"]
    price = []
    for i in range(2, sheet1.max_row + 1):
        price.append(sheet1[f"N{i}"].value)
    return price

# create a list of both lists (dates and ratios) for the javascript 'Chart.js'

def createList(date, price):
    i = 0
    data = []
    while i < len(date):
        data.append((date[i], price[i]))
        i = i+1
    return data


def plotyChart(filename, stock1, stock2, curr1, curr2):
    df = pd.read_excel(filename, sheet_name="Sheet1")
    fig = go.Figure(data=
                    [go.Candlestick(x=df["Dates"],
                                    open=df["Open"],
                                    high=df["High"],
                                    low=df["Low"],
                                    close=df["Close"])]
                    )

    fig.update_layout(
        title=f"{stock1} priced in {stock2}",
        yaxis_title=f"{stock1} ({curr1}) / {stock2} ({curr2})"
    )

    return fig
