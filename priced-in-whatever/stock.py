
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import yfinance as yf
import pandas as pd
import time
from Calender import *
from date import *
import os



### Dates ##############################################################################################################

# create a table (xlsx file), just to compare the start dates and see the the entered dates are valid.

def tableForDates(stock1, stock2):
    tickers = [stock1, stock2]
    interval = '1d'
    period1 = int(time.mktime(datetime.datetime(1971, 1, 1, 23, 59).timetuple()))
    period2 = int(time.mktime(datetime.datetime(2022, 1, 30, 23, 59).timetuple()))

    filename = f'{stock1}pricedin{stock2}.xlsx'
    xlwriter = pd.ExcelWriter(filename, engine='openpyxl')

    for ticker in tickers:
        query_string = f'https://query1.finance.yahoo.com/v7/finance/download/{ticker}?period1={period1}&period2={period2}&interval={interval}&events=history&includeAdjustedClose=true'
        df = pd.read_csv(query_string)
        df.to_excel(xlwriter, sheet_name=ticker, index=False)

    xlwriter.save()


# get first open value (when closing is missing)

def openValue(filename, stock):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[f"{stock}"]
    return sheet["B2"].value

def emptyStart(filename, stock1, stock2):
    wb = openpyxl.load_workbook("pricedinwhatever.xlsx")
    dateSheet = wb["Sheet1"]
    ind = 2

    while dateSheet[f"B{ind}"].value == None or dateSheet[f"B{ind}"].value == 0:
        dateSheet[f"B{ind}"].value = openValue(filename, stock1)
        ind = ind + 1

    ind = 2
    while dateSheet[f"C{ind}"].value == None or dateSheet[f"C{ind}"].value == 0:
        dateSheet[f"C{ind}"].value = openValue(filename, stock2)
        ind = ind + 1

    wb.save("pricedinwhatever.xlsx")

# verify not blank dates

def removeBlanks():
    wb = openpyxl.load_workbook("pricedinwhatever.xlsx")
    dateSheet = wb["Sheet1"]

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"B{i}"].value == None or dateSheet[f"B{i}"].value == 0:
            dateSheet[f"B{i}"].value = dateSheet[f"B{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"C{i}"].value == None or dateSheet[f"C{i}"].value == 0:
            dateSheet[f"C{i}"].value = dateSheet[f"C{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"D{i}"].value == None or dateSheet[f"D{i}"].value == 0:
            dateSheet[f"D{i}"].value = dateSheet[f"D{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"E{i}"].value == None or dateSheet[f"E{i}"].value == 0:
            dateSheet[f"E{i}"].value = dateSheet[f"E{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"F{i}"].value == None or dateSheet[f"F{i}"].value == 0:
            dateSheet[f"F{i}"].value = dateSheet[f"F{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"G{i}"].value == None or dateSheet[f"G{i}"].value == 0:
            dateSheet[f"G{i}"].value = dateSheet[f"G{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"H{i}"].value == None or dateSheet[f"H{i}"].value == 0:
            dateSheet[f"H{i}"].value = dateSheet[f"H{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"I{i}"].value == None or dateSheet[f"I{i}"].value == 0:
            dateSheet[f"I{i}"].value = dateSheet[f"I{i-1}"].value
    wb.save("pricedinwhatever.xlsx")


### Calculate ##########################################################################################################

# create a table (xlsx file) with two stocks and the start and end dates, calculate the ratio between both stocks.

def createTable(stock1, stock2, date1, date2):
    print(date1)
    print(date2)
    tickers = [stock1, stock2]
    interval = '1d'
    period1 = int(time.mktime(datetime.datetime(getYear(date1), getMonth(date1), getDay(date1), 23, 59).timetuple()))
    period2 = int(time.mktime(datetime.datetime(getYear(date2), getMonth(date2), getDay(date2), 23, 59).timetuple()))

    filename = f'{stock1}pricedin{stock2}.xlsx'
    xlwriter = pd.ExcelWriter(filename, engine='openpyxl')

    for ticker in tickers:
        query_string = f'https://query1.finance.yahoo.com/v7/finance/download/{ticker}?period1={period1}&period2={period2}&interval={interval}&events=history&includeAdjustedClose=true'
        df = pd.read_csv(query_string)
        df.to_excel(xlwriter, sheet_name=ticker, index=False)
    xlwriter.save()


# calculate the ratio between two stocks

def calculate(filename):
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb["Sheet1"]
    sheet1["K1"].value = "Open"
    sheet1["L1"].value = "High"
    sheet1["M1"].value = "Low"
    sheet1["N1"].value = "Close"
    for i in range(2, sheet1.max_row + 1):
        sheet1[f"K{i}"].value = sheet1[f"B{i}"].value / sheet1[f"F{i}"].value
    for i in range(2, sheet1.max_row + 1):
        sheet1[f"L{i}"].value = sheet1[f"C{i}"].value / sheet1[f"G{i}"].value
    for i in range(2, sheet1.max_row + 1):
        sheet1[f"M{i}"].value = sheet1[f"D{i}"].value / sheet1[f"H{i}"].value
    for i in range(2, sheet1.max_row + 1):
        sheet1[f"N{i}"].value = sheet1[f"E{i}"].value / sheet1[f"I{i}"].value
    wb.save(filename)

    """

    sheet1.column_dimensions['A'].width = 20
    sheet1.column_dimensions['B'].width = 20
    sheet1.column_dimensions['C'].width = 20
    sheet1.column_dimensions['D'].width = 20
    sheet1.column_dimensions['E'].width = 20
    sheet1.column_dimensions['F'].width = 20

    sheet1['A1'].font = Font(bold=True)
    sheet1['B1'].font = Font(bold=True)
    sheet1['C1'].font = Font(bold=True)
    sheet1['D1'].font = Font(bold=True)
    sheet1['E1'].font = Font(bold=True)
    sheet1['F1'].font = Font(bold=True)
    """


# convert oz to gram

def ozToGram(obj):
    return obj/31.1035


# get the first date a stock was traded

def getFirstDate(filename, stock):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[stock]
    return sheet["A2"].value

# get the first common date of two stocks

def getCommonDate(filename, stock1, stock2):
    tableForDates(stock1, stock2)
    date1 = cleanDate(getFirstDate(filename, stock1))
    date2 = cleanDate(getFirstDate(filename, stock2))
    if compare(date1,date2) == 1:
        return date1
    else:
        return date2

######## fix dates ########

def fixDates(dateStart, dateEnd):
    dates = pd.date_range(start=dateStart, end=dateEnd)
    df = dates.strftime("%Y-%m-%d")
    df = df.to_frame("dates")
    df.to_excel("pricedinwhatever.xlsx")

def fillList(filename, stock1, stock2):
    ind = 2

    wb1 = openpyxl.load_workbook("pricedinwhatever.xlsx")
    dateSheet = wb1["Sheet1"]

    wb2 = openpyxl.load_workbook(filename)
    stockSheet = wb2[stock1]


    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"A{i}"].value == stockSheet[f"A{ind}"].value:
            dateSheet[f"B{i}"].value = stockSheet[f"B{ind}"].value
            dateSheet[f"C{i}"].value = stockSheet[f"B{ind}"].value
            dateSheet[f"D{i}"].value = stockSheet[f"B{ind}"].value
            dateSheet[f"E{i}"].value = stockSheet[f"B{ind}"].value
            ind = ind + 1
        else:
            dateSheet[f"B{i}"].value = dateSheet[f"B{i-1}"].value  # copy the date of the date before
            dateSheet[f"C{i}"].value = stockSheet[f"C{ind-1}"].value
            dateSheet[f"D{i}"].value = stockSheet[f"D{ind-1}"].value
            dateSheet[f"E{i}"].value = stockSheet[f"E{ind-1}"].value

    ind = 2
    stockSheet = wb2[stock2]

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"A{i}"].value == stockSheet[f"A{ind}"].value:
            dateSheet[f"F{i}"].value = stockSheet[f"B{ind}"].value
            dateSheet[f"G{i}"].value = stockSheet[f"C{ind}"].value
            dateSheet[f"H{i}"].value = stockSheet[f"D{ind}"].value
            dateSheet[f"I{i}"].value = stockSheet[f"E{ind}"].value
            ind = ind + 1
        else:
            dateSheet[f"C{i}"].value = dateSheet[f"C{i-1}"].value  # copy the date of the date before

    dateSheet["A1"].value = "Dates"
    dateSheet["B1"].value = f"{stock1} - Open"
    dateSheet["C1"].value = f"{stock1} - High"
    dateSheet["D1"].value = f"{stock1} - Low"
    dateSheet["E1"].value = f"{stock1} - Close"
    dateSheet["F1"].value = f"{stock2} - Open"
    dateSheet["G1"].value = f"{stock2} - High"
    dateSheet["H1"].value = f"{stock2} - Low"
    dateSheet["I1"].value = f"{stock2} - Close"


    wb1.save("pricedinwhatever.xlsx")


### for 2 stocks and a conversion ###

# convert the currency of stock (will occur when both stocks are priced by the same currency)

def convertRate(stock1, stock2, converisonRate, date1, date2):
    tickers = [stock1, stock2, converisonRate]
    interval = '1d'
    period1 = int(time.mktime(datetime.datetime(getYear(date1), getMonth(date1), getDay(date1), 23, 59).timetuple()))
    period2 = int(time.mktime(datetime.datetime(getYear(date2), getMonth(date2), getDay(date2), 23, 59).timetuple()))

    filename = f'{stock1}pricedin{stock2}.xlsx'
    xlwriter = pd.ExcelWriter(filename, engine='openpyxl')

    for ticker in tickers:
        query_string = f'https://query1.finance.yahoo.com/v7/finance/download/{ticker}?period1={period1}&period2={period2}&interval={interval}&events=history&includeAdjustedClose=true'
        df = pd.read_csv(query_string)
        df.to_excel(xlwriter, sheet_name=ticker, index=False)
    xlwriter.save()

# remove blanks in table

def removeBlanksConver():
    wb = openpyxl.load_workbook("pricedinwhatever.xlsx")
    dateSheet = wb["Sheet1"]

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"B{i}"].value == None or dateSheet[f"B{i}"].value == 0:
            dateSheet[f"B{i}"].value = dateSheet[f"B{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"C{i}"].value == None or dateSheet[f"C{i}"].value == 0:
            dateSheet[f"C{i}"].value = dateSheet[f"C{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"D{i}"].value == None or dateSheet[f"D{i}"].value == 0:
            dateSheet[f"D{i}"].value = dateSheet[f"D{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"E{i}"].value == None or dateSheet[f"E{i}"].value == 0:
            dateSheet[f"E{i}"].value = dateSheet[f"E{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"F{i}"].value == None or dateSheet[f"F{i}"].value == 0:
            dateSheet[f"F{i}"].value = dateSheet[f"F{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"G{i}"].value == None or dateSheet[f"G{i}"].value == 0:
            dateSheet[f"G{i}"].value = dateSheet[f"G{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"H{i}"].value == None or dateSheet[f"H{i}"].value == 0:
            dateSheet[f"H{i}"].value = dateSheet[f"H{i-1}"].value

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"I{i}"].value == None or dateSheet[f"I{i}"].value == 0:
            dateSheet[f"I{i}"].value = dateSheet[f"I{i-1}"].value
    wb.save("pricedinwhatever.xlsx")

# calculate the ratio between two stocks and conversion rate

def calculateConver(filename):
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb["Sheet1"]
    sheet1["K1"].value = "Open"
    sheet1["L1"].value = "High"
    sheet1["M1"].value = "Low"
    sheet1["N1"].value = "Close"
    for i in range(2, sheet1.max_row+1):
        sheet1[f"K{i}"].value = sheet1[f"B{i}"].value / (sheet1[f"F{i}"].value / sheet1[f"J{i}"].value)
    for i in range(2, sheet1.max_row + 1):
        sheet1[f"L{i}"].value = sheet1[f"C{i}"].value / (sheet1[f"G{i}"].value / sheet1[f"J{i}"].value)
    for i in range(2, sheet1.max_row+1):
        sheet1[f"M{i}"].value = sheet1[f"D{i}"].value / (sheet1[f"H{i}"].value / sheet1[f"J{i}"].value)
    for i in range(2, sheet1.max_row+1):
        sheet1[f"N{i}"].value = sheet1[f"E{i}"].value / (sheet1[f"I{i}"].value / sheet1[f"J{i}"].value)
    wb.save(filename)

    """
    
    sheet1.column_dimensions['A'].width = 20
    sheet1.column_dimensions['B'].width = 20
    sheet1.column_dimensions['C'].width = 20
    sheet1.column_dimensions['D'].width = 20
    sheet1.column_dimensions['E'].width = 20
    sheet1.column_dimensions['F'].width = 20

    sheet1['A1'].font = Font(bold=True)
    sheet1['B1'].font = Font(bold=True)
    sheet1['C1'].font = Font(bold=True)
    sheet1['D1'].font = Font(bold=True)
    sheet1['E1'].font = Font(bold=True)
    sheet1['F1'].font = Font(bold=True)
    """



def fillListConver(filename, stock1, stock2, conversionRate):

    fillList(filename, stock1, stock2)

    wb1 = openpyxl.load_workbook("pricedinwhatever.xlsx")
    dateSheet = wb1["Sheet1"]
    wb2 = openpyxl.load_workbook(filename)
    stockSheet = wb2[conversionRate]

    for i in range(2, dateSheet.max_row + 1):
        if dateSheet[f"A{i}"].value == stockSheet[f"A{ind}"].value:
            dateSheet[f"J{i}"].value = stockSheet[f"E{ind}"].value
            ind = ind + 1
        else:
            dateSheet[f"J{i}"].value = dateSheet[f"J{i-1}"].value  # copy the date of the date before

    dateSheet["J1"].value = conversionRate

    wb1.save("pricedinwhatever.xlsx")
